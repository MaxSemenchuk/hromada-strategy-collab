# DEPRECATED: use KSE-Loc-Data-Hub partnerships-hromadas-network.csv instead
# (see data/sources/kse-pin.json and docs/kse-synergy.md). Kept for research-log provenance.
import openpyxl, re, json
import datetime

wb = openpyxl.load_workbook('mss_registry.xlsx', read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))[1:]
rows = [r for r in rows if r and r[0] is not None]

# Filter to rows added 2021-2022 (post-reform, modern hromada names expected)
recent = []
for r in rows:
    num, title, date_added, date_num, subjects, responsible, form, term, changes = (list(r) + [None]*9)[:9]
    if isinstance(date_added, datetime.datetime) and date_added.year in (2021, 2022):
        recent.append({'num': num, 'title': title, 'date': date_added, 'subjects': subjects or '', 'form': form or ''})

print(f"2021-2022 rows: {len(recent)}")

# Better party extractor: find ALL "Name + (сільська|селищна|міська) рада" patterns
PARTY_RE = re.compile(
    r"([А-ЯҐЄІЇа-яґєії'’\-]+(?:ська|цька|зька))\s+(сільська|селищна|міська)\s+рад[аи]",
    re.UNICODE
)

def category(title, form):
    t = (str(title) + ' ' + str(form)).lower()
    if 'адміністративних послуг' in t: return 'Адмінпослуги'
    if 'пожежн' in t: return 'Пожежна охорона'
    if 'архітектурно-будівельного' in t: return 'Архбудконтроль'
    if 'відход' in t or 'сміт' in t or 'тпв' in t: return 'Відходи'
    if 'водопостачан' in t or 'водовідведен' in t: return 'Вода'
    if 'спільного комунального підприємства' in t: return 'Спільне КП'
    if 'спільного фінансування' in t or 'утримання' in t: return 'Спільне фінансування установ'
    if 'туристич' in t: return 'Туризм'
    if 'освіт' in t or 'школ' in t: return 'Освіта'
    if 'медичн' in t or 'охорони здоров' in t: return 'Медицина'
    if 'дорог' in t or 'шлях' in t or 'міст' in t: return 'Дороги'
    if 'реєстрац' in t: return 'Реєстрація актів'
    return 'Спільний проект (інше)'

edges = []
node_names = set()
for r in recent:
    parties_raw = PARTY_RE.findall(r['subjects'])
    parties = list(dict.fromkeys([f"{name}{' ' if not name.endswith(chr(8217)) else ''}{typ}" for name, typ in parties_raw]))
    if len(parties) < 2:
        continue
    cat = category(r['title'], r['form'])
    for i in range(len(parties)):
        for j in range(i+1, len(parties)):
            edges.append({'a': parties[i], 'b': parties[j], 'category': cat, 'date': r['date'].strftime('%Y-%m-%d'), 'title': str(r['title'])[:80]})
            node_names.add(parties[i]); node_names.add(parties[j])

print(f"Rows with 2+ parties extracted: {len(set((e['a'],e['b'],e['date']) for e in edges))}")
print(f"Unique nodes (hromadas): {len(node_names)}")
print(f"Total edges: {len(edges)}")

from collections import Counter
cat_counts = Counter(e['category'] for e in edges)
print("\nCategories in this slice:")
for c, n in cat_counts.most_common():
    print(f"  {c}: {n}")

print("\nSample edges:")
for e in edges[:15]:
    print(f"  {e['a']} <-> {e['b']}  [{e['category']}]  {e['date']}")

json.dump({'nodes': sorted(node_names), 'edges': edges}, open('mss_graph_mvp_data.json','w'), ensure_ascii=False, indent=2)
