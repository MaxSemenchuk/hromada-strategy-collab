import openpyxl, re
from collections import Counter, defaultdict

wb = openpyxl.load_workbook('mss_registry.xlsx', read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))[1:]  # skip header
rows = [r for r in rows if r and r[0] is not None]
print(f"Total rows: {len(rows)}")

# --- Categorize by form of cooperation (col index 6 = 'Форма співробітництва', but let's also scan title col 1) ---
def categorize(title, form_text):
    t = (str(title) + ' ' + str(form_text)).lower()
    if 'адміністративних послуг' in t or 'адмін' in t and 'послуг' in t:
        return 'Адміністративні послуги (ЦНАП)'
    if 'пожежн' in t:
        return 'Пожежна охорона'
    if 'архітектурно-будівельного контролю' in t or 'архбуд' in t:
        return 'Архітектурно-будівельний контроль'
    if 'реєстрац' in t and ('акт' in t or 'громадян' in t or 'державн' in t):
        return 'Державна реєстрація актів'
    if 'відход' in t or 'сміт' in t or 'тпв' in t:
        return 'Поводження з відходами'
    if 'водопостачан' in t or 'водовідведен' in t or 'водоканал' in t:
        return 'Водопостачання/водовідведення'
    if 'спільного комунального підприємства' in t or 'утворення' in t and 'підприємств' in t:
        return 'Створення спільного КП'
    if 'спільного фінансування' in t or 'утримання' in t:
        return 'Спільне фінансування/утримання установ'
    if 'туристич' in t:
        return 'Туризм'
    if 'освіт' in t or 'школ' in t or 'днз' in t or 'дошкільн' in t:
        return 'Освіта'
    if 'медичн' in t or 'охорони здоров' in t or 'амбулатор' in t:
        return 'Охорона здоров\'я'
    if 'дорог' in t or 'шлях' in t or 'вулиц' in t or 'мост' in t or 'міст' in t and 'ремонт' in t:
        return 'Дороги/інфраструктура'
    if 'спільного проекту' in t or 'спільних проектів' in t or 'спільного проєкту' in t:
        return 'Спільний проект (інше)'
    if 'делегування' in t:
        return 'Делегування завдань (інше)'
    return 'Інше/не визначено'

cat_counter = Counter()
year_counter = Counter()
party_counter = Counter()
n_parties_dist = Counter()
oblast_counter = Counter()

# regex to find oblast mentions
oblast_re = re.compile(r'([А-ЯҐЄІЇ][а-яґєії\'’-]+(?:ому|ої|ій|ку|ка|ий|а)?\s+(?:обл|область))', re.UNICODE)

for r in rows:
    num, title, date_added, date_num, subjects, responsible, form, term, changes = (list(r) + [None]*9)[:9]
    cat = categorize(title, form)
    cat_counter[cat] += 1
    if isinstance(date_added, __import__('datetime').datetime):
        year_counter[date_added.year] += 1
    subj_text = str(subjects or '')
    # rough split of parties by known council-type suffixes
    parties = re.split(r'(?<=[а-яіїєґ])\s+(?=[А-ЯҐЄІЇ][а-яіїєґ]+(?:ська|цька|зька)\s+(?:сільська|селищна|міська)\s+рада)', subj_text)
    parties = [p.strip() for p in parties if len(p.strip()) > 5]
    n_parties_dist[min(len(parties), 6)] += 1
    for p in parties:
        m = re.match(r"([А-ЯҐЄІЇ][а-яіїєґ'’-]+(?:ська|цька|зька))\s+(сільська|селищна|міська)\s+рада", p)
        if m:
            party_counter[m.group(1) + ' ' + m.group(2)] += 1
    for om in oblast_re.findall(subj_text):
        oblast_counter[om] += 1

print("\n=== ФОРМИ СПІВРОБІТНИЦТВА (топ) ===")
for cat, cnt in cat_counter.most_common(20):
    print(f"  {cat:<45} {cnt:>4}  ({100*cnt/len(rows):.1f}%)")

print("\n=== РОЗПОДІЛ ЗА РОКОМ ВНЕСЕННЯ ДО РЕЄСТРУ ===")
for y, c in sorted(year_counter.items()):
    print(f"  {y}: {c}")

print("\n=== СКІЛЬКИ СТОРІН У ДОГОВОРІ (розподіл) ===")
for n, c in sorted(n_parties_dist.items()):
    label = f"{n}+" if n == 6 else str(n)
    print(f"  {label} сторін: {c} договорів")

print("\n=== ТОП-20 ГРОМАД ЗА КІЛЬКІСТЮ УГОД (найактивніші учасники МСС) ===")
for name, cnt in party_counter.most_common(20):
    print(f"  {name:<35} {cnt} угод")

print(f"\nUnique parties identified: {len(party_counter)}")
